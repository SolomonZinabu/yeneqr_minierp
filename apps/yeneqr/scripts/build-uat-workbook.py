"""
YeneQR UAT Test Case Workbook
=============================
End-to-end test cases for User Acceptance Testing.
One sheet per module + a Summary dashboard with pass/fail counters.

Modules covered:
  1. Customer QR Flow
  2. Orders + Kitchen
  3. Menu Management
  4. Tables + Floors
  5. Staff + Shifts
  6. Settings + Branches
  7. Promotions + Loyalty
  8. Analytics + CRM
  9. Reviews + Reservations + Waitlist
 10. Localization + i18n

Each test case row has columns:
  Test ID | Module | Severity | Test Name | Preconditions | Steps | Expected Result | Test Data | Status | Tester | Date | Bug Link | Actual Result

Output: /home/z/my-project/download/YeneQR_UAT_TestCases.xlsx
"""

import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ============================================================
# Design tokens
# ============================================================
PRIMARY = "1B2A4A"
PRIMARY_LIGHT = "D6E4F0"
ACCENT_POSITIVE = "1B7D46"   # green = pass
ACCENT_NEGATIVE = "C0392B"   # red = fail
ACCENT_WARNING = "D4820A"    # amber = blocked
NEUTRAL_900 = "37352F"
NEUTRAL_600 = "8C8A84"
NEUTRAL_200 = "E9E9E8"
NEUTRAL_100 = "F7F7F5"
WHITE = "FFFFFF"

FONT_NAME = "Noto Sans CJK SC"  # Linux fallback — works for English too

thin = Side(border_style="thin", color=NEUTRAL_200)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============================================================
# Column layout for test case sheets
# ============================================================
COLUMNS = [
    ("Test ID",          12),
    ("Severity",          9),
    ("Test Name",        32),
    ("Preconditions",    38),
    ("Steps",            60),
    ("Expected Result",  45),
    ("Test Data",        28),
    ("Status",           12),
    ("Tester",           14),
    ("Date",             12),
    ("Bug Link",         22),
    ("Actual Result",    32),
]

HEADER_ROW = 4
FIRST_DATA_ROW = 5

# ============================================================
# Test case data
# Each module is a list of (severity, name, preconditions, steps, expected, test_data) tuples
# Severity: P0 (blocker), P1 (critical), P2 (major), P3 (minor)
# ============================================================

MODULES = {}

# ============================================================
# 1. CUSTOMER QR FLOW
# ============================================================
MODULES["1. Customer QR Flow"] = [
    ("P0", "Scan QR → load menu",
     "QR code generated for Table T-01, Bole branch. Restaurant active. Menu has items with images.",
     "1. Open phone camera\n2. Scan QR code on Table T-01\n3. Wait for redirect to menu URL",
     "Menu page loads within 3s. Restaurant name, logo, and menu categories visible. No console errors.",
     "QR for Table T-01, Bole"),

    ("P0", "Add item to cart",
     "Menu page loaded. At least one available item.",
     "1. Tap an available menu item\n2. Tap 'Add to Cart'\n3. Check cart icon badge count",
     "Item added to cart. Cart badge shows '1'. Item name, price, and quantity visible in cart.", ""),

    ("P0", "Add item with modifiers",
     "Menu item has at least one modifier group (e.g., Size: Small/Large).",
     "1. Tap item with modifiers\n2. Select 'Large' size\n3. Tap 'Add to Cart'\n4. Open cart",
     "Cart shows item with 'Large' modifier. Price reflects modifier delta (e.g., +50 ETB).", ""),

    ("P0", "Remove ingredients from item",
     "Menu item has linked ingredients (e.g., Doro Wot with onion, garlic).",
     "1. Tap item\n2. Tap 'Remove' on onion\n3. Add to cart\n4. Open cart item details",
     "Cart item shows 'Remove: onion' in special instructions. Order POST includes removedIngredients.", ""),

    ("P0", "Place order (dine-in, cash)",
     "Cart has ≥1 item. Table T-01 scanned. Customer session active.",
     "1. Open cart\n2. Tap 'Place Order'\n3. Select 'Cash' payment\n4. Confirm",
     "Order created with status 'pending'. Order number shown (e.g., #0001). Order tracking screen appears. Kitchen view receives new order.", ""),

    ("P0", "Place order (takeaway)",
     "Cart has ≥1 item. Customer on takeaway menu (no table).",
     "1. Open cart\n2. Select 'Takeaway' order type\n3. Tap 'Place Order'",
     "Order created with type='takeaway'. No tableId in order. Packaging fee applied if configured.", ""),

    ("P1", "Apply promotion code at checkout",
     "Active promotion with code 'WELCOME10' (10% off).",
     "1. Add items to cart (subtotal ≥ minimum order)\n2. Enter code 'WELCOME10'\n3. Tap 'Apply'",
     "Discount applied. Cart total reduced by 10%. Promo code shown in order summary.", "Code: WELCOME10"),

    ("P1", "Redeem loyalty points",
     "Customer has ≥100 loyalty points. Loyalty enabled for restaurant.",
     "1. Add items to cart\n2. Toggle 'Use loyalty points'\n3. Place order",
     "Loyalty points deducted. Cart total reduced by points × pointValue. Points balance updated after order.", ""),

    ("P1", "Call waiter from menu",
     "Customer on menu page. Waiter calls enabled in restaurant settings.",
     "1. Tap 'Call Waiter' button\n2. Wait for acknowledgment",
     "Waiter view shows new call with table number. Status: 'pending' → 'acknowledged' when waiter taps it.", ""),

    ("P1", "View order status tracking",
     "Order placed successfully.",
     "1. After placing order, stay on tracking screen\n2. Watch status changes as kitchen processes",
     "Status updates in real-time: pending → accepted → preparing → ready → served. Each transition visible without refresh.", ""),

    ("P1", "Place second round on same order",
     "First order placed, status='accepted' or 'preparing'.",
     "1. Tap 'Add more items'\n2. Add new item to cart\n3. Tap 'Place Order' (round 2)",
     "Second round added to same order. Round number increments to 2. Kitchen view shows new items with same order number.", ""),

    ("P2", "Add special instructions to item",
     "Item in cart.",
     "1. Tap item in cart\n2. Add 'Extra spicy, no onion' in instructions\n3. Save",
     "Instructions saved on cart item. Sent to kitchen in order item. Visible in KDS item card.", ""),

    ("P2", "Favorite / unfavorite item",
     "Customer logged in (phone number entered).",
     "1. Tap heart icon on a menu item\n2. Refresh page\n3. Tap heart again",
     "Heart fills when favorited. State persists across refresh. Unfavoriting removes it.", ""),

    ("P2", "Switch language (EN ↔ AM)",
     "Restaurant has Amharic translations enabled.",
     "1. Tap language switcher\n2. Select 'አማርኺ'\n3. Browse menu",
     "UI strings translate to Amharic. Menu item names show Amharic if available. Text is left-to-right (NOT RTL).", ""),

    ("P2", "Leave review after order",
     "Order status='completed'. Reviews enabled.",
     "1. On order completion screen, tap 'Leave Review'\n2. Select 4 stars\n3. Write review text\n4. Submit",
     "Review saved. Appears in restaurant's Reviews tab. Owner can reply.", ""),

    ("P3", "Cart persists across page refresh",
     "Cart has items. Customer on menu page.",
     "1. Add items to cart\n2. Refresh browser\n3. Open cart",
     "Cart items persist after refresh (localStorage).", ""),

    # Negative tests
    ("P1", "Cannot order when restaurant closed",
     "Restaurant working hours set to 02:00-06:00 (currently 10:00).",
     "1. Scan QR\n2. Try to place order",
     "Order blocked with message 'Restaurant is currently closed'. Cart disabled or warning shown.", ""),

    ("P2", "Cannot add unavailable item",
     "Item marked isAvailable=false.",
     "1. Browse menu\n2. Find unavailable item\n3. Try to add to cart",
     "Item shows 'Unavailable' badge. Add-to-cart button disabled or hidden.", ""),

    ("P2", "Promo code rejection (invalid code)",
     "Cart has items.",
     "1. Enter code 'INVALID123'\n2. Tap 'Apply'",
     "Error message: 'Invalid or expired promotion code'. No discount applied.", "Code: INVALID123"),

    ("P2", "Promo code rejection (minimum order not met)",
     "Promo 'WELCOME10' has minimumOrder=500 ETB. Cart subtotal=200 ETB.",
     "1. Add items totaling 200 ETB\n2. Enter 'WELCOME10'\n3. Apply",
     "Error: 'Minimum order of 500 ETB required for this promotion'.", "Subtotal: 200 ETB"),

    ("P2", "Duplicate order prevention",
     "Order just placed from table T-01 within last 30s.",
     "1. Immediately try to place another order from same table",
     "Warning: 'An order was just placed for this table. Please wait a moment.' Order blocked.", ""),
]

# ============================================================
# 2. ORDERS + KITCHEN
# ============================================================
MODULES["2. Orders + Kitchen"] = [
    ("P0", "Order list loads with branch filter",
     "Logged in as owner. Branch 'Bole' selected.",
     "1. Navigate to Orders\n2. Select 'Bole' branch filter\n3. Check order list",
     "Only orders from Bole branch shown. No orders from CMC branch leak through.", ""),

    ("P0", "Accept pending order (waiter_first mode)",
     "Order routing = 'waiter_first'. New order exists with status='pending'.",
     "1. Open Orders view\n2. Find pending order\n3. Click 'Accept'\n4. Verify status change",
     "Order status changes pending → accepted. Order disappears from 'Pending' filter, appears in 'Accepted'.", ""),

    ("P0", "Direct-to-kitchen auto-accept",
     "Order routing = 'direct_to_kitchen'. Customer places new order.",
     "1. As customer, place order\n2. As staff, open Orders view\n3. Check new order status",
     "Order created directly in 'accepted' status (skips 'pending'). Order events timeline shows system auto-accept with AUTO badge.", ""),

    ("P0", "Kitchen Display System (KDS) shows new order",
     "Order with status='accepted' or 'preparing'.",
     "1. Open Kitchen view\n2. Check new order appears",
     "Order card visible in KDS with order number, table, items, and timers. New order chime plays.", ""),

    ("P0", "Mark item as preparing (KDS)",
     "Order in KDS with items in 'pending' kitchen status.",
     "1. Tap 'Start Preparing' on an item\n2. Verify status",
     "Item kitchen status: pending → preparing. Timer starts. Item card color changes.", ""),

    ("P0", "Mark item as ready (KDS)",
     "Item in 'preparing' status.",
     "1. Tap 'Mark Ready' on the item",
     "Item status: preparing → ready. Waiter view notified. Order ready sound plays in waiter view.", ""),

    ("P1", "Mark all items ready at once",
     "Order has multiple items in 'preparing' status.",
     "1. Tap 'Mark All Ready' on the order",
     "All items transition to 'ready'. Order status auto-advances to 'ready' if all items ready.", ""),

    ("P1", "Order status transitions (full lifecycle)",
     "New pending order.",
     "1. Accept (pending → accepted)\n2. Start preparing (accepted → preparing)\n3. Mark ready (preparing → ready)\n4. Pick up (ready → picked_up)\n5. Serve (picked_up → served)\n6. Mark paid (served → paid)\n7. Complete (paid → completed)",
     "Each transition succeeds. Order events timeline records every change with actor and timestamp.", ""),

    ("P1", "Cancel order with reason",
     "Order in 'pending' or 'accepted' status.",
     "1. Open order detail\n2. Click 'Cancel Order'\n3. Enter reason 'Customer changed mind'\n4. Confirm",
     "Order status → 'cancelled'. Reason saved in order record. Table freed (if dine-in). Cancellation event in timeline.", ""),

    ("P1", "Order Events Timeline",
     "Order with at least 3 status changes.",
     "1. Open order detail sheet\n2. Click 'Timeline' button in header",
     "Dialog opens showing chronological list of events. Each event shows: type, from→to status, actor (customer/staff/system), timestamp. AUTO badge on system auto-accepts.", ""),

    ("P1", "Waiter view filters by assigned waiter",
     "Waiter 'Abebe' assigned to tables T-01, T-02. Orders exist on both.",
     "1. Log in as Abebe\n2. Open Waiter view",
     "Only orders for T-01 and T-02 shown. Orders on T-03 (assigned to other waiter) not visible.", ""),

    ("P1", "Split bill",
     "Order with status='served', multiple items.",
     "1. Open order detail\n2. Click 'Split Bill'\n3. Create 2 splits (e.g., 2 items each)\n4. Save",
     "Bill splits created. Each split has its own subtotal. Payment can be made per split.", ""),

    ("P1", "Payment (cash) marks order paid",
     "Order status='served'.",
     "1. Open order detail\n2. Click 'Mark Paid' (Cash)\n3. Confirm amount",
     "Payment recorded. Order status → 'paid'. Receipt printable. Order moves to 'Completed' tab after.", ""),

    ("P2", "VIP priority order badge",
     "Order placed from VIP table (isVip=true).",
     "1. Place order from VIP table\n2. Open Orders view",
     "Order shows '⭐ VIP' badge. Order sorted higher in kitchen view.", ""),

    ("P2", "Rush priority order",
     "Order with priority='rush'.",
     "1. As staff, mark order as 'Rush'\n2. Check KDS",
     "Order shows '🔥 RUSH' badge. Sorted to top of KDS queue.", ""),

    ("P2", "Print kitchen ticket",
     "Order in KDS.",
     "1. Click 'Print Kitchen Ticket' on an order",
     "Print dialog opens with formatted kitchen ticket (order #, table, items, modifiers, special instructions).", ""),

    ("P2", "Print receipt",
     "Order status='paid' or 'completed'.",
     "1. Open order detail\n2. Click 'Print Receipt'",
     "Receipt preview opens with restaurant info, order items, totals, payment method, change due.", ""),

    ("P3", "Order search by order number",
     "Multiple orders exist.",
     "1. Enter '#0001' in search box\n2. Verify filter",
     "Only order #0001 shown in list.", "Search: #0001"),

    # Negative
    ("P1", "Invalid status transition blocked",
     "Order in 'pending' status.",
     "1. Try to mark order as 'completed' directly (skip accepted/preparing/ready)",
     "Transition blocked with error: 'Cannot transition from pending to completed'. State machine enforced.", ""),

    ("P2", "86'd item (kitchen cancel)",
     "Order item in 'preparing' status.",
     "1. In KDS, tap '86' on an item\n2. Confirm cancellation",
     "Item status → 'cancelled'. Order total recalculated (item removed). Customer notified if applicable.", ""),
]

# ============================================================
# 3. MENU MANAGEMENT
# ============================================================
MODULES["3. Menu Management"] = [
    ("P0", "Create menu",
     "Logged in as owner or manager.",
     "1. Go to Menu view\n2. Click 'Add Menu'\n3. Enter name 'Lunch Menu'\n4. Save",
     "Menu created. Appears in menu list. Can be assigned to QR codes.", "Menu: Lunch Menu"),

    ("P0", "Create category in menu",
     "Menu 'Lunch Menu' exists.",
     "1. Select 'Lunch Menu'\n2. Click 'Add Category'\n3. Enter name 'Starters'\n4. Save",
     "Category created under 'Lunch Menu'. Visible in menu structure.", "Category: Starters"),

    ("P0", "Create menu item with image",
     "Category 'Starters' exists.",
     "1. Click 'Add Item' in Starters\n2. Enter name 'Tibs'\n3. Set price 350 ETB\n4. Upload image\n5. Save",
     "Item created with image. Image persists after page refresh. Item visible in customer menu with image.", "Item: Tibs, 350 ETB"),

    ("P0", "Edit menu item price",
     "Item 'Tibs' exists at 350 ETB.",
     "1. Edit item 'Tibs'\n2. Change price to 400 ETB\n3. Save",
     "Price updated. Customer menu shows 400 ETB. Existing orders retain old price.", ""),

    ("P0", "Mark item unavailable",
     "Item 'Tibs' exists.",
     "1. Edit item\n2. Toggle 'Available' to off\n3. Save",
     "Item marked unavailable. Customer menu shows 'Unavailable' badge. Cannot be added to cart.", ""),

    ("P1", "Add modifier group to item",
     "Item 'Tibs' exists.",
     "1. Edit item\n2. Add modifier group 'Size'\n3. Add options: Small (+0), Large (+50)\n4. Save",
     "Modifier group saved. Customer can select size when adding to cart. Price delta applied.", ""),

    ("P1", "Link ingredients to item",
     "Ingredients 'onion', 'garlic', 'berbere' exist.",
     "1. Edit item\n2. Go to Ingredients tab\n3. Link onion (removable=true, default=true)\n4. Save",
     "Ingredients linked. Customer can remove 'onion' from cart item. KDS shows ingredient list.", ""),

    ("P1", "Set item availability schedule",
     "Item 'Tibs' exists.",
     "1. Edit item\n2. Set 'Available 09:00-14:00, weekdays only'\n3. Save",
     "Schedule saved. Item only appears in customer menu during those hours on weekdays. After 14:00, marked unavailable.", ""),

    ("P1", "Create combo meal",
     "Multiple items exist.",
     "1. Create new item 'Combo Special'\n2. Toggle 'This is a combo meal'\n3. Add 2 included items (e.g., Tibs + Rice)\n4. Save",
     "Combo saved. Customer sees combo with included items listed. Price uses combo price, not sum of items.", ""),

    ("P2", "Item branch override (price)",
     "Restaurant has 2 branches (Bole, CMC). Item 'Tibs' default price 350.",
     "1. Edit item\n2. Add branch override for CMC: price=400\n3. Save",
     "Bole shows 350, CMC shows 400. Orders at each branch use the correct price.", ""),

    ("P2", "Reorder categories (drag-drop)",
     "Multiple categories exist.",
     "1. Drag 'Desserts' above 'Mains'\n2. Save order",
     "Category order updated. Customer menu reflects new order.", ""),

    ("P2", "Bulk menu import",
     "Menu JSON file prepared.",
     "1. Click 'Bulk Import'\n2. Upload JSON file\n3. Confirm import",
     "Items/categories created from JSON. Existing items not duplicated (matched by name).", ""),

    ("P3", "Item description with Amharic",
     "Item 'Tibs' exists.",
     "1. Edit item\n2. Add Amharic description 'ጥብስ በበርበሬ'\n3. Save",
     "Amharic description saved. Renders LTR (not RTL) in customer menu when Amharic selected.", ""),

    # Negative
    ("P1", "Cannot delete item with active orders",
     "Item 'Tibs' has an order in 'pending' status.",
     "1. Try to delete item 'Tibs'",
     "Deletion blocked. Error: 'Cannot delete item with active orders'. Soft-delete (mark unavailable) suggested.", ""),

    ("P2", "Modifier option with invalid price",
     "Modifier group 'Size' exists.",
     "1. Add option 'Large' with price -50\n2. Save",
     "Validation error: 'Price cannot be negative'. Option not saved.", ""),

    ("P2", "Item name uniqueness per category",
     "Item 'Tibs' exists in 'Starters'.",
     "1. Add new item in 'Starters'\n2. Name it 'Tibs'\n3. Try to save",
     "Warning or error about duplicate name. Either block or allow with confirmation.", ""),
]

# ============================================================
# 4. TABLES + FLOORS
# ============================================================
MODULES["4. Tables + Floors"] = [
    ("P0", "Create floor plan",
     "Branch 'Bole' selected.",
     "1. Go to Tables view\n2. Click 'Add Floor'\n3. Name it 'Ground Floor'\n4. Save",
     "Floor created. Floor tab appears. Canvas editor opens for the new floor.", ""),

    ("P0", "Add table to floor",
     "Floor 'Ground Floor' exists.",
     "1. Select 'Ground Floor' tab\n2. Click 'Add Table'\n3. Enter number 'T-01', capacity 4\n4. Save",
     "Table T-01 added to canvas. Visible on floor plan. QR code auto-generated.", ""),

    ("P0", "Edit table (capacity, shape)",
     "Table T-01 exists.",
     "1. Double-click T-01\n2. Change capacity to 6, shape to 'rectangle'\n3. Save",
     "Table updated. Shape changes on canvas. Capacity shows in table detail.", ""),

    ("P0", "Generate QR code for table",
     "Table T-01 exists, no QR yet.",
     "1. Select T-01\n2. Click 'Generate QR'\n3. Download/print",
     "QR code generated with URL pointing to menu for T-01's branch. QR downloadable as PNG.", ""),

    ("P1", "Merge two tables",
     "Tables T-01 and T-02 both have active orders.",
     "1. Click 'Merge Tables'\n2. Select T-01 as primary\n3. Select T-02 as secondary\n4. Confirm",
     "Items from T-02's order transferred to T-01's order. T-02 marked 'available'. T-02's original order cancelled.", ""),

    ("P1", "Table status auto-updates",
     "Table T-01 status='available'.",
     "1. Customer at T-01 places order\n2. Check Tables view",
     "Table T-01 status auto-changes to 'occupied'. Real-time update without refresh.", ""),

    ("P1", "Free table after order completes",
     "Table T-01 occupied, order status='completed'.",
     "1. Mark table as 'available' (or auto-free on completion)\n2. Verify status",
     "Table status → 'available'. Ready for next customer.", ""),

    ("P2", "VIP table flag",
     "Table T-01 exists.",
     "1. Edit T-01\n2. Toggle 'VIP table'\n3. Save",
     "VIP flag saved. Orders from T-01 get priority='vip'. Special badge in orders view.", ""),

    ("P2", "Assign specific menu to table",
     "Multiple menus exist (e.g., 'Standard', 'VIP Menu').",
     "1. Edit T-01\n2. Set menuId='VIP Menu'\n3. Save\n4. Scan QR for T-01",
     "Customer sees 'VIP Menu' when scanning T-01's QR, not the default menu.", ""),

    ("P2", "Floor plan wall drawing",
     "Floor 'Ground Floor' in edit mode.",
     "1. Select 'Draw Wall' tool\n2. Click and drag to draw a wall segment\n3. Save",
     "Wall segment appears on floor plan. Persists after save. Visible to staff viewing layout.", ""),

    ("P3", "Table notes",
     "Table T-01 exists.",
     "1. Edit T-01\n2. Add note 'Window seat, near outlet'\n3. Save",
     "Note saved. Visible to staff in table detail. Helps with seating decisions.", ""),

    # Negative
    ("P1", "Cannot delete table with active order",
     "Table T-01 has order in 'pending' status.",
     "1. Try to delete T-01",
     "Deletion blocked. Error: 'Cannot delete table with active orders'.", ""),

    ("P2", "Merge tables from different branches",
     "T-01 in Bole, T-05 in CMC.",
     "1. Click 'Merge Tables'\n2. Select T-01 (primary)\n3. Try to select T-05 (secondary)",
     "T-05 not available in secondary dropdown (filtered to same branch). Cannot merge across branches.", ""),
]

# ============================================================
# 5. STAFF + SHIFTS
# ============================================================
MODULES["5. Staff + Shifts"] = [
    ("P0", "Create staff member (waiter)",
     "Logged in as owner.",
     "1. Go to Staff view\n2. Click 'Add Staff'\n3. Enter name, email, role='waiter'\n4. Assign to Bole branch\n5. Save",
     "Staff created. Receives login credentials. Appears in staff list with 'waiter' role.", ""),

    ("P0", "Staff login",
     "Staff account created with password.",
     "1. Log out\n2. Go to login page\n3. Enter staff email + password\n4. Log in",
     "Staff logged in. Sees only permitted views (no Settings if no permission). Branch scope enforced.", ""),

    ("P0", "Assign waiter to tables",
     "Waiter 'Abebe' exists. Tables T-01, T-02 exist.",
     "1. Go to Staff view → Waiter Assignments\n2. Assign Abebe to T-01 and T-02\n3. Save",
     "Assignment saved. Abebe sees orders from T-01, T-02 in Waiter view. Other waiters don't see them.", ""),

    ("P1", "Edit staff permissions",
     "Staff member 'cashier1' exists.",
     "1. Edit cashier1\n2. Add permission 'order:manage'\n3. Remove 'menu:manage'\n4. Save",
     "Permissions updated. cashier1 can now manage orders but not edit menu. Enforced on next API call.", ""),

    ("P1", "Deactivate staff",
     "Active staff member exists.",
     "1. Edit staff\n2. Toggle 'Active' to off\n3. Save",
     "Staff deactivated. Cannot log in. Existing shift entries preserved.", ""),

    ("P1", "Start shift",
     "Staff member active, no open shift.",
     "1. As staff, click 'Start Shift'\n2. Confirm",
     "Shift entry created with startTime=now. Active shift shown in Shifts view.", ""),

    ("P1", "End shift",
     "Staff has active shift.",
     "1. Click 'End Shift'\n2. Confirm",
     "Shift entry closed with endTime=now. Duration calculated. Shift history updated.", ""),

    ("P2", "Staff avatar upload",
     "Staff member exists.",
     "1. Edit staff\n2. Upload avatar photo\n3. Save",
     "Avatar saved (was previously dropped on create — now fixed). Visible in staff list and waiter view.", ""),

    ("P2", "Branch-scoped staff list",
     "Manager logged in, assigned to Bole branch only.",
     "1. Go to Staff view",
     "Only Bole branch staff visible. CMC staff not shown (branch scope enforced).", ""),

    ("P2", "Role-based UI restrictions",
     "Logged in as kitchen_staff role.",
     "1. Navigate sidebar",
     "Kitchen view visible. Orders view visible. Settings, Promotions, Staff management hidden (no permission).", ""),

    ("P3", "Staff password reset",
     "Staff member exists.",
     "1. As owner, edit staff\n2. Click 'Reset Password'\n3. Enter new password\n4. Save",
     "Password updated. Staff can log in with new password. Old password rejected.", ""),

    # Negative
    ("P1", "Duplicate email rejected",
     "Staff with email 'abebe@yeneqr.com' exists.",
     "1. Add new staff\n2. Use email 'abebe@yeneqr.com'\n3. Try to save",
     "Error: 'A staff member with this email already exists'. Save blocked.", "Email: abebe@yeneqr.com"),

    ("P2", "Weak password rejected",
     "Creating new staff.",
     "1. Enter password '123'\n2. Save",
     "Validation error: 'Password must be at least 8 characters'. Save blocked.", "Password: 123"),
]

# ============================================================
# 6. SETTINGS + BRANCHES
# ============================================================
MODULES["6. Settings + Branches"] = [
    ("P0", "Update restaurant profile",
     "Logged in as owner.",
     "1. Settings → Profile\n2. Update name, description, cuisine type\n3. Save",
     "Profile saved. Changes visible on customer landing page and dashboard.", ""),

    ("P0", "Set working hours",
     "Restaurant profile open.",
     "1. Settings → Working Hours\n2. Set Mon-Fri 08:00-22:00, Sat-Sun 10:00-23:00\n3. Save",
     "Hours saved. Customer menu shows 'Open/Closed' based on current time. Orders blocked when closed.", ""),

    ("P0", "Configure tax rate and service charge",
     "Settings → Tax & Service.",
     "1. Set taxRate=0.15 (15%)\n2. Set serviceCharge=0.10 (10%)\n3. Save",
     "Tax and service charge saved. Applied to all new orders. Visible in cart totals.", ""),

    ("P0", "Enable/disable payment methods",
     "Settings → Payment Methods.",
     "1. Enable 'Cash', 'Telebirr'\n2. Disable 'Chapa'\n3. Save",
     "Only enabled methods shown at customer checkout. Disabled methods hidden.", ""),

    ("P0", "Order routing flag (restaurant-level)",
     "Settings → Feature Settings → Kitchen.",
     "1. Set 'Order Routing' = 'Direct to Kitchen'\n2. Save\n3. Place test order",
     "New orders auto-accept to 'accepted' status. KDS receives immediately. Waiter still assigned for delivery.", ""),

    ("P0", "Order routing flag (branch override)",
     "Restaurant routing='direct_to_kitchen'. Branch 'CMC' exists.",
     "1. Settings → Branch Settings\n2. Select CMC\n3. Set orderRouting='waiter_first'\n4. Save\n5. Place order at CMC",
     "CMC orders land in 'pending' (waiter_first wins over restaurant default). Bole still uses direct_to_kitchen.", ""),

    ("P1", "Create new branch",
     "Logged in as owner.",
     "1. Go to Branches (sidebar or Settings)\n2. Click 'Add Branch'\n3. Enter name 'Piassa', address\n4. Save",
     "Branch created. Appears in branch switcher. Can assign staff, tables, menus to it.", ""),

    ("P1", "Branch-level tax override",
     "Branch 'CMC' exists. Restaurant taxRate=0.15.",
     "1. Settings → Branch Settings → CMC\n2. Set taxRate=0.10\n3. Save",
     "CMC orders use 10% tax. Bole still uses 15%. Branch override wins.", ""),

    ("P1", "StarPay configuration",
     "Settings → StarPay tab.",
     "1. Enter StarPay merchant ID, API secret, webhook secret\n2. Save\n3. Test payment",
     "StarPay credentials saved. Test payment initiates StarPay flow. Webhook updates payment status.", ""),

    ("P1", "AI provider configuration",
     "Settings → AI Configuration.",
     "1. Select provider='gemini'\n2. Enter API key\n3. Select model='gemini-1.5-flash'\n4. Enable AI chat\n5. Save",
     "AI config saved. Customer chat widget activates. AI suggestions work in dashboard.", ""),

    ("P2", "Switch active branch",
     "Owner with access to multiple branches.",
     "1. Click branch switcher in sidebar\n2. Select 'CMC'",
     "All views update to show CMC data only. Orders, tables, staff filtered to CMC.", ""),

    ("P2", "Subscription panel",
     "Subscription exists for restaurant.",
     "1. Settings → Subscription tab",
     "Current plan, status, billing period visible. Recent invoices listed. Change/cancel plan buttons work.", ""),

    ("P2", "Dynamic pricing rule",
     "Settings → Dynamic Pricing tab.",
     "1. Click 'New Rule'\n2. Name 'Afternoon Happy Hour'\n3. Set 20% off, 14:00-17:00 weekdays\n4. Save",
     "Rule created. Shows 'LIVE' badge during active window. Items discounted in customer menu during those hours.", ""),

    ("P2", "Branch settings reset to inherit",
     "Branch 'CMC' has taxRate=0.10 override.",
     "1. Branch Settings → CMC\n2. Click 'Inherit' on taxRate\n3. Save",
     "Override cleared. CMC now uses restaurant default (0.15).", ""),

    ("P3", "2FA setup",
     "Owner account.",
     "1. Settings → Security\n2. Enable 2FA\n3. Scan QR with authenticator app\n4. Enter code\n5. Verify",
     "2FA enabled. Future logins require code from authenticator. Backup codes provided.", ""),

    # Negative
    ("P1", "Invalid AI API key",
     "AI config open.",
     "1. Enter invalid API key\n2. Save\n3. Try AI chat",
     "Config saved but AI chat returns error: 'AI provider authentication failed'. Falls back gracefully.", ""),

    ("P2", "StarPay webhook signature mismatch",
     "StarPay configured with wrong webhook secret.",
     "1. Simulate webhook with wrong signature\n2. Check API response",
     "Webhook rejected with 401. Payment status NOT updated. Logs show signature mismatch.", ""),
]

# ============================================================
# 7. PROMOTIONS + LOYALTY
# ============================================================
MODULES["7. Promotions + Loyalty"] = [
    ("P0", "Create percentage promotion",
     "Logged in as owner.",
     "1. Go to Promotions\n2. Click 'Create'\n3. Name 'Summer 15%'\n4. Type=discount, percentage=15\n5. Valid dates\n6. Save",
     "Promotion created. Visible in promotions list. Customer can apply if eligible.", ""),

    ("P0", "Create coupon code promotion",
     "Promotions view.",
     "1. Click 'Create'\n2. Type=coupon\n3. Code='SUMMER15'\n4. 15% off\n5. Save",
     "Coupon created with code. Customer must enter code at checkout to get discount.", "Code: SUMMER15"),

    ("P0", "Apply promotion in customer cart",
     "Active promotion 'SUMMER15' (15% off, no minimum).",
     "1. Add items to cart (subtotal 500 ETB)\n2. Enter 'SUMMER15'\n3. Apply",
     "Discount of 75 ETB applied. New total: 425 ETB + tax. Promo shown in order summary.", "Subtotal: 500 ETB"),

    ("P0", "Create loyalty reward (discount)",
     "Loyalty enabled.",
     "1. Settings → Loyalty\n2. Add reward\n3. Name='50 ETB off'\n4. Type=discount, amount=50\n5. Points cost=500\n6. Save",
     "Reward created. Visible in customer loyalty rewards list. Customer can redeem at 500 points.", ""),

    ("P0", "Create loyalty reward (free item)",
     "Menu item 'Tibs' exists. Loyalty enabled.",
     "1. Add reward\n2. Type='Free Item'\n3. Select 'Tibs' from dropdown\n4. Points cost=1000\n5. Save",
     "Reward created with linked menu item. Customer can redeem at 1000 points for free Tibs.", ""),

    ("P1", "Customer earns loyalty points",
     "Customer has phone number. Loyalty enabled. pointsPerAmount=10 (10 ETB = 1 point).",
     "1. Customer places order, subtotal 500 ETB\n2. Check customer loyalty balance",
     "Customer earns 50 points. Balance updated. Visible in customer profile and CRM.", ""),

    ("P1", "Redeem loyalty reward",
     "Customer has 1000 points. Free item reward exists.",
     "1. Customer opens rewards\n2. Tap 'Redeem' on free item\n3. Confirm",
     "1000 points deducted. Free item added to cart (or order). Points balance updated.", ""),

    ("P1", "Deactivate promotion",
     "Active promotion exists.",
     "1. Go to Promotions\n2. Toggle 'Active' to off",
     "Promotion deactivated. No longer applicable at checkout. Existing discounted orders keep their discount.", ""),

    ("P2", "Promotion with usage limit",
     "Promotion 'SUMMER15' with usageLimit=100.",
     "1. Apply promotion 100 times\n2. Try 101st application",
     "After 100 uses, promotion rejected: 'Usage limit reached'. 101st attempt blocked.", ""),

    ("P2", "Promotion branch assignment",
     "Promotion 'SUMMER15' created. Branches Bole, CMC exist.",
     "1. Edit promotion\n2. Assign to Bole only\n3. Save\n4. Try at CMC",
     "Promotion works at Bole. Rejected at CMC: 'Not valid at this branch'.", ""),

    ("P2", "Happy hour auto-applied",
     "Happy hour promotion active (14:00-17:00, 20% off).",
     "1. At 15:00, add items to cart\n2. Check totals",
     "20% discount auto-applied (no code needed). Visible in order summary. After 17:00, no discount.", ""),

    ("P3", "Promotion Amharic name",
     "Creating promotion.",
     "1. Enter Amharic name 'የበጋ ቅናሽ'\n2. Save\n3. Switch customer to Amharic",
     "Amharic name displays LTR (not RTL). Persists after save.", ""),

    # Negative
    ("P1", "Percentage > 100 rejected",
     "Creating promotion.",
     "1. Type=percentage, value=150\n2. Save",
     "Validation error: 'Percentage cannot exceed 100%'. Save blocked.", ""),

    ("P2", "Coupon code uniqueness",
     "Coupon 'SUMMER15' exists.",
     "1. Create new coupon with code 'SUMMER15'\n2. Save",
     "Error: 'A promotion with this code already exists'. Save blocked.", ""),

    ("P2", "Free item reward without menuItemId",
     "Creating loyalty reward.",
     "1. Type='free_item'\n2. Don't select menu item\n3. Try to save",
     "Save button disabled or error: 'Please select a menu item'. Cannot create free_item without menuItemId.", ""),
]

# ============================================================
# 8. ANALYTICS + CRM
# ============================================================
MODULES["8. Analytics + CRM"] = [
    ("P0", "Dashboard shows today's stats",
     "Orders exist for today.",
     "1. Open Dashboard",
     "Today's revenue, order count, avg order value, busy hours visible. Numbers match orders view.", ""),

    ("P0", "Analytics view renders charts",
     "Orders exist over past 7 days.",
     "1. Go to Analytics view\n2. Select 'Last 7 days' range",
     "Revenue chart, orders chart, top items chart render without errors. Data matches dashboard.", ""),

    ("P0", "Export analytics to CSV",
     "Analytics view loaded with data.",
     "1. Click 'Export'\n2. Download CSV",
     "CSV downloads with order-level data. Columns: date, order number, total, items, payment method.", ""),

    ("P1", "CRM customer list",
     "Customers exist (placed orders with phone number).",
     "1. Go to CRM view",
     "Customer list shows: name, phone, visit count, total spent, last visit. Sortable by total spent.", ""),

    ("P1", "Customer detail view",
     "Customer exists with order history.",
     "1. Click customer in CRM\n2. View detail",
     "Shows customer info, loyalty points, order history, favorite items, total spent.", ""),

    ("P1", "Menu engineering matrix",
     "Menu items with order history exist.",
     "1. Go to Menu Engineering view",
     "Items classified as Star/Puzzle/Plowhorse/Dog based on popularity vs margin. Quadrant chart renders.", ""),

    ("P2", "Analytics date range filter",
     "Data exists for multiple months.",
     "1. Select date range 'Last 30 days'\n2. Then 'Last 7 days'\n3. Compare",
     "Charts update to reflect selected range. Numbers change appropriately.", ""),

    ("P2", "Top-selling items report",
     "Orders exist.",
     "1. Analytics → Top Items tab",
     "Ranked list of items by quantity sold and revenue. Shows item name, count, revenue, % of total.", ""),

    ("P2", "Customer search in CRM",
     "Multiple customers exist.",
     "1. Enter phone number in search\n2. Filter",
     "Customer with matching phone shown. Partial match works.", "Search: 0911"),

    ("P3", "CRM customer segments",
     "Customers with varying visit counts.",
     "1. CRM → Segments tab",
     "Customers auto-segmented: New, Regular, VIP, At-risk. Counts shown per segment.", ""),

    # Negative
    ("P2", "Analytics with no data",
     "New restaurant, no orders yet.",
     "1. Open Analytics",
     "Empty state shown: 'No data yet'. No chart errors. Encouraging message displayed.", ""),
]

# ============================================================
# 9. REVIEWS + RESERVATIONS + WAITLIST
# ============================================================
MODULES["9. Reviews + Reservations + Waitlist"] = [
    ("P0", "Customer leaves review",
     "Order status='completed'. Reviews enabled.",
     "1. On order completion, tap 'Review'\n2. 5 stars, 'Great food!'\n3. Submit",
     "Review saved. Visible in Reviews view for staff.", ""),

    ("P0", "Owner replies to review",
     "Review exists without owner reply.",
     "1. Go to Reviews view\n2. Click 'Reply' on a review\n3. Enter 'Thank you!'\n4. Send",
     "Reply saved. Visible on customer-facing review display (if applicable). Reply timestamp recorded.", ""),

    ("P0", "Create reservation",
     "Reservations enabled. Tables exist.",
     "1. Go to Reservations view\n2. Click 'New Reservation'\n3. Enter customer name, phone, party size=4, date/time\n4. Save",
     "Reservation created. Table held for that time. Appears in reservations list.", ""),

    ("P0", "Waitlist add customer",
     "Waitlist enabled. Restaurant busy.",
     "1. Go to Waitlist view\n2. Click 'Add to Waitlist'\n3. Enter name, phone, party size=2\n4. Save",
     "Customer added to waitlist. Position shown. Estimated wait time calculated.", ""),

    ("P1", "Waitlist seat customer",
     "Customer in waitlist, table becomes available.",
     "1. Click 'Seat' on waitlist entry\n2. Select available table\n3. Confirm",
     "Customer removed from waitlist. Table marked occupied. Notification sent to customer.", ""),

    ("P1", "Reservation status transitions",
     "Reservation exists with status='pending'.",
     "1. Confirm reservation (pending → confirmed)\n2. On arrival, seat (confirmed → seated)\n3. After order, complete (seated → completed)",
     "Each transition works. Table status syncs with reservation status.", ""),

    ("P1", "Review rating filter",
     "Reviews with mixed ratings exist.",
     "1. Reviews view\n2. Filter by '1-2 stars'",
     "Only low-rated reviews shown. Helps owner spot issues.", ""),

    ("P2", "Waitlist SMS notification",
     "Waitlist entry with phone number. SMS configured.",
     "1. Customer's turn approaching\n2. Click 'Notify Customer'",
     "SMS sent to customer: 'Your table is ready at [restaurant]'. Logged in notifications.", ""),

    ("P2", "Reservation no-show",
     "Reservation confirmed, customer doesn't arrive.",
     "1. After 15 min grace period, mark as 'no-show'\n2. Confirm",
     "Reservation status → 'no_show'. Table freed for other customers. Customer flagged in CRM.", ""),

    ("P2", "Waitlist auto-removal on timeout",
     "Waitlist entry with estimated wait of 30 min.",
     "1. Wait 30+ min without seating\n2. Check waitlist",
     "Entry auto-marked as 'expired' or remains with 'late' indicator. Staff can call customer.", ""),

    ("P3", "Review with photo",
     "Customer completed order.",
     "1. Leave review\n2. Attach photo of food\n3. Submit",
     "Photo uploaded. Visible in review. Thumbnail in reviews list.", ""),

    # Negative
    ("P1", "Double-booking table reservation",
     "Table T-01 reserved for 19:00-21:00.",
     "1. Try to reserve T-01 for 20:00 same day",
     "Error: 'Table not available for this time slot'. Suggests alternative tables or times.", ""),

    ("P2", "Review on uncompleted order",
     "Order status='preparing' (not completed).",
     "1. Try to leave review",
     "Review option not available. Message: 'Available after order is completed'.", ""),
]

# ============================================================
# 10. LOCALIZATION + i18n
# ============================================================
MODULES["10. Localization + i18n"] = [
    ("P0", "Enable Amharic language",
     "Restaurant has i18n access.",
     "1. Go to Localization view\n2. Languages tab\n3. Enable 'Amharic'\n4. Save",
     "Amharic enabled. Customer can switch to Amharic. Default language unchanged (English).", ""),

    ("P0", "Edit UI string translation",
     "Amharic enabled.",
     "1. Localization → UI Strings tab\n2. Find 'Add to Cart'\n3. Edit Amharic translation to 'ወደ ጋሪ ጨምር'\n4. Save",
     "Translation saved. Customer menu in Amharic shows the new string.", ""),

    ("P0", "Bulk AI translate menu items",
     "AI configured (gemini/openai). Amharic enabled. Menu items exist.",
     "1. Localization → Menu Content tab\n2. Click 'Bulk Translate'\n3. Select Amharic\n4. Run",
     "Items translated. Names and descriptions in Amharic. Toast shows count translated/skipped.", ""),

    ("P1", "Customer language switcher",
     "Amharic enabled with translations.",
     "1. Open customer menu\n2. Tap language switcher\n3. Select አማርኺ",
     "UI switches to Amharic. Menu names show Amharic where available. Text renders LTR (not RTL).", ""),

    ("P1", "Menu content manual translation",
     "Menu item 'Tibs' exists.",
     "1. Localization → Menu Content\n2. Find 'Tibs'\n3. Enter Amharic name 'ጥብስ'\n4. Save",
     "Translation saved. Customer in Amharic sees 'ጥብስ' instead of 'Tibs'.", ""),

    ("P2", "Disable language",
     "Amharic enabled.",
     "1. Localization → Languages\n2. Disable Amharic\n3. Save",
     "Amharic removed from customer language switcher. Existing translations preserved in DB.", ""),

    ("P2", "Translation overview cards",
     "Multiple languages enabled with varying translation coverage.",
     "1. Localization → Overview tab",
     "Cards show per-language completion % (e.g., Amharic 75%, Oromo 30%). Helps prioritize translation work.", ""),

    ("P2", "Bulk translate with overwrite",
     "Items already have Amharic translations.",
     "1. Bulk Translate\n2. Toggle 'Overwrite existing'\n3. Run",
     "Existing translations replaced with new AI translations. Count shows 'translated' not 'skipped'.", ""),

    ("P3", "Add custom language",
     "Localization → Languages.",
     "1. Click 'Add Language'\n2. Enter code 'om' (Oromo)\n3. Set direction='ltr'\n4. Save",
     "Language added. Appears in switcher. Translations can now be entered for Oromo.", ""),

    # Negative
    ("P1", "Bulk translate without AI configured",
     "AI not configured (provider='none').",
     "1. Bulk Translate\n2. Run",
     "Error: 'AI not configured. Set up AI in Settings first.' No translations made.", ""),

    ("P2", "Invalid language code",
     "Adding new language.",
     "1. Enter code 'xyz' (not a real language)\n2. Save",
     "Warning or validation error. Either blocked or allowed with confirmation.", ""),
]

# ============================================================
# Build the workbook
# ============================================================
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# Helper: create a styled sheet for a module
def create_module_sheet(name, cases):
    ws = wb.create_sheet(title=name[:31])  # Excel sheet name max 31 chars

    # Title row (row 2)
    ws.cell(row=2, column=2, value=f"YeneQR UAT — {name}").font = Font(
        name=FONT_NAME, size=14, bold=True, color=PRIMARY
    )
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(COLUMNS)+1)

    # Subtitle (row 3)
    ws.cell(row=3, column=2, value=f"{len(cases)} test cases • Severity: P0 (blocker) / P1 (critical) / P2 (major) / P3 (minor)").font = Font(
        name=FONT_NAME, size=9, italic=True, color=NEUTRAL_600
    )
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(COLUMNS)+1)

    # Header row (row 4)
    header_fill = PatternFill("solid", fgColor=PRIMARY)
    header_font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=2):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[HEADER_ROW].height = 32

    # Data rows
    data_font = Font(name=FONT_NAME, size=9, color=NEUTRAL_900)
    alt_fill = PatternFill("solid", fgColor=NEUTRAL_100)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for idx, (severity, test_name, preconditions, steps, expected, test_data) in enumerate(cases):
        row = FIRST_DATA_ROW + idx
        test_id = f"{name.split('.')[0]}.T{idx+1:03d}"

        values = [
            test_id,
            severity,
            test_name,
            preconditions,
            steps,
            expected,
            test_data,
            "Not Run",   # Status
            "",          # Tester
            "",          # Date
            "",          # Bug Link
            "",          # Actual Result
        ]

        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = BORDER_ALL
            if idx % 2 == 1:
                cell.fill = alt_fill

        # Color severity cell
        sev_cell = ws.cell(row=row, column=3)  # Severity is col 3 (B=2, C=3)
        sev_colors = {
            "P0": ACCENT_NEGATIVE,
            "P1": ACCENT_WARNING,
            "P2": "F4D03F",  # yellow
            "P3": NEUTRAL_600,
        }
        sev_cell.font = Font(name=FONT_NAME, size=9, bold=True, color=sev_colors.get(severity, NEUTRAL_900))
        sev_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Set row height based on content (rough heuristic)
        max_lines = max(
            len(str(preconditions)) // 40 + 1,
            len(str(steps)) // 55 + 1,
            len(str(expected)) // 40 + 1,
        )
        ws.row_dimensions[row].height = max(30, min(120, max_lines * 14))

    # Status dropdown (data validation)
    status_col_letter = get_column_letter(2 + 7)  # Status is 8th column → col 9 → letter I
    status_range = f"{status_col_letter}{FIRST_DATA_ROW}:{status_col_letter}{FIRST_DATA_ROW + len(cases) - 1}"
    dv = DataValidation(
        type="list",
        formula1='"Not Run,Pass,Fail,Blocked,N/A"',
        allow_blank=True,
        showDropDown=False,  # show dropdown arrow
    )
    dv.add(status_range)
    ws.add_data_validation(dv)

    # Conditional formatting on Status column
    pass_fill = PatternFill("solid", fgColor="D4EDDA")  # light green
    fail_fill = PatternFill("solid", fgColor="F8D7DA")  # light red
    blocked_fill = PatternFill("solid", fgColor="FFF3CD")  # light amber
    notrun_fill = PatternFill("solid", fgColor=NEUTRAL_100)

    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Pass"'], fill=pass_fill, font=Font(name=FONT_NAME, size=9, color=ACCENT_POSITIVE))
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Fail"'], fill=fail_fill, font=Font(name=FONT_NAME, size=9, color=ACCENT_NEGATIVE))
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=blocked_fill, font=Font(name=FONT_NAME, size=9, color=ACCENT_WARNING))
    )

    # Freeze panes: header row + first 3 columns (Test ID, Severity, Test Name)
    ws.freeze_panes = "E5"

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"

    return ws, len(cases)

# Create all module sheets
module_stats = []
for module_name, cases in MODULES.items():
    ws, count = create_module_sheet(module_name, cases)
    module_stats.append((module_name, count))

# ============================================================
# Summary Dashboard sheet (insert at position 0)
# ============================================================
summary = wb.create_sheet(title="Summary", index=0)

# Title
summary.cell(row=2, column=2, value="YeneQR — UAT Test Case Summary").font = Font(
    name=FONT_NAME, size=16, bold=True, color=PRIMARY
)
summary.merge_cells("B2:G2")

summary.cell(row=3, column=2, value="User Acceptance Testing • Production launch readiness").font = Font(
    name=FONT_NAME, size=10, italic=True, color=NEUTRAL_600
)
summary.merge_cells("B3:G3")

# Test data info
summary.cell(row=5, column=2, value="Test Environment:").font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
summary.cell(row=5, column=3, value="https://yeneqr.techbee.et").font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
summary.cell(row=6, column=2, value="Test Restaurant:").font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
summary.cell(row=6, column=3, value="Habesha Maebel (slug: habesha-maebel)").font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
summary.cell(row=7, column=2, value="Owner Login:").font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
summary.cell(row=7, column=3, value="owner@habeshamaebel.com / admin123").font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
summary.cell(row=8, column=2, value="Branches:").font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
summary.cell(row=8, column=3, value="Bole (main), CMC").font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)

# Module progress table
table_start_row = 11
headers = ["Module", "Total Cases", "Pass", "Fail", "Blocked", "Not Run", "Progress %"]
for col_idx, h in enumerate(headers, start=2):
    cell = summary.cell(row=table_start_row, column=col_idx, value=h)
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER_ALL

# Compute sheet name → safe reference for COUNTIF
def safe_sheet_name(name):
    """Excel sheet reference: single quotes if name has spaces or special chars."""
    if any(c in name for c in " -'"):
        return f"'{name[:31]}'"
    return name[:31]

for idx, (module_name, count) in enumerate(module_stats):
    row = table_start_row + 1 + idx
    sheet_ref = safe_sheet_name(module_name)
    status_col = "I"  # Status is column I (9th, starting from A=1, B=2...)
    data_start = 5
    data_end = 4 + count

    summary.cell(row=row, column=2, value=module_name).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    summary.cell(row=row, column=3, value=count).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    # Pass count
    summary.cell(row=row, column=4, value=f'=COUNTIF({sheet_ref}!{status_col}{data_start}:{status_col}{data_end},"Pass")').font = Font(name=FONT_NAME, size=10, color=ACCENT_POSITIVE)
    # Fail count
    summary.cell(row=row, column=5, value=f'=COUNTIF({sheet_ref}!{status_col}{data_start}:{status_col}{data_end},"Fail")').font = Font(name=FONT_NAME, size=10, color=ACCENT_NEGATIVE)
    # Blocked count
    summary.cell(row=row, column=6, value=f'=COUNTIF({sheet_ref}!{status_col}{data_start}:{status_col}{data_end},"Blocked")').font = Font(name=FONT_NAME, size=10, color=ACCENT_WARNING)
    # Not Run = Total - Pass - Fail - Blocked
    summary.cell(row=row, column=7, value=f'={count}-D{row}-E{row}-F{row}').font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600)
    # Progress % = (Pass + Fail) / Total  (i.e., executed)
    summary.cell(row=row, column=8, value=f'=IFERROR((D{row}+E{row})/C{row},0)').number_format = "0%"
    summary.cell(row=row, column=8).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)

    for col_idx in range(2, 9):
        summary.cell(row=row, column=col_idx).border = BORDER_ALL
        if idx % 2 == 1:
            summary.cell(row=row, column=col_idx).fill = PatternFill("solid", fgColor=NEUTRAL_100)
        summary.cell(row=row, column=col_idx).alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")

# Totals row
totals_row = table_start_row + 1 + len(module_stats)
summary.cell(row=totals_row, column=2, value="TOTAL").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
summary.cell(row=totals_row, column=3, value=f"=SUM(C{table_start_row+1}:C{totals_row-1})").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
summary.cell(row=totals_row, column=4, value=f"=SUM(D{table_start_row+1}:D{totals_row-1})").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
summary.cell(row=totals_row, column=5, value=f"=SUM(E{table_start_row+1}:E{totals_row-1})").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
summary.cell(row=totals_row, column=6, value=f"=SUM(F{table_start_row+1}:F{totals_row-1})").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
summary.cell(row=totals_row, column=7, value=f"=SUM(G{table_start_row+1}:G{totals_row-1})").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
summary.cell(row=totals_row, column=8, value=f"=IFERROR((D{totals_row}+E{totals_row})/C{totals_row},0)").font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
summary.cell(row=totals_row, column=8).number_format = "0%"

for col_idx in range(2, 9):
    summary.cell(row=totals_row, column=col_idx).fill = PatternFill("solid", fgColor=PRIMARY)
    summary.cell(row=totals_row, column=col_idx).border = BORDER_ALL
    summary.cell(row=totals_row, column=col_idx).alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")

# Column widths for summary
summary_widths = [2, 32, 12, 10, 10, 12, 12, 14]
for col_idx, w in enumerate(summary_widths, start=1):
    summary.column_dimensions[get_column_letter(col_idx)].width = w

# Legend
legend_row = totals_row + 3
summary.cell(row=legend_row, column=2, value="Severity Legend:").font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
legend_items = [
    ("P0", "Blocker — demo-breaking, must fix before launch", ACCENT_NEGATIVE),
    ("P1", "Critical — major feature broken, fix before launch", ACCENT_WARNING),
    ("P2", "Major — works but with issues, fix within 1 week", "F4D03F"),
    ("P3", "Minor — cosmetic or nice-to-have, fix when possible", NEUTRAL_600),
]
for i, (sev, desc, color) in enumerate(legend_items):
    r = legend_row + 1 + i
    summary.cell(row=r, column=2, value=sev).font = Font(name=FONT_NAME, size=10, bold=True, color=color)
    summary.cell(row=r, column=3, value=desc).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    summary.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

# Status legend
status_legend_row = legend_row + len(legend_items) + 3
summary.cell(row=status_legend_row, column=2, value="Status Legend:").font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
status_items = [
    ("Pass", "Test executed, expected result achieved"),
    ("Fail", "Test executed, expected result NOT achieved — log a bug"),
    ("Blocked", "Cannot execute (dependency missing, env issue) — log blocker"),
    ("Not Run", "Not yet executed (default state)"),
    ("N/A", "Not applicable to this environment"),
]
for i, (status, desc) in enumerate(status_items):
    r = status_legend_row + 1 + i
    summary.cell(row=r, column=2, value=status).font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
    summary.cell(row=r, column=3, value=desc).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    summary.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

# Instructions
instr_row = status_legend_row + len(status_items) + 3
summary.cell(row=instr_row, column=2, value="How to use:").font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
instructions = [
    "1. Assign each module to a tester (add name in 'Tester' column on each module sheet).",
    "2. Execute test cases in order — P0 first (blockers), then P1, P2, P3.",
    "3. For each case: set Status dropdown (Pass/Fail/Blocked/N/A), enter Date, add Bug Link if failed.",
    "4. If Fail: write what actually happened in 'Actual Result' column. Log bug in GitHub Issues or tracker.",
    "5. Summary tab auto-updates with pass/fail counts per module via COUNTIF formulas.",
    "6. Launch readiness: all P0 must Pass. P1 ≥ 95% Pass. P2 ≥ 80% Pass. P3 at discretion.",
]
for i, instr in enumerate(instructions):
    r = instr_row + 1 + i
    summary.cell(row=r, column=2, value=instr).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    summary.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

summary.freeze_panes = "B11"

# Workbook metadata
wb.properties.creator = "Z.ai"
wb.properties.title = "YeneQR UAT Test Cases"
wb.properties.subject = "User Acceptance Testing — Pre-Launch"

# Save
output_path = "/home/z/my-project/download/YeneQR_UAT_TestCases.xlsx"
wb.save(output_path)

# Report
total_cases = sum(c for _, c in module_stats)
print(f"✅ Workbook saved: {output_path}")
print(f"   Sheets: {len(wb.sheetnames)} ({len(MODULES)} modules + 1 Summary)")
print(f"   Total test cases: {total_cases}")
print(f"   Breakdown by module:")
for name, count in module_stats:
    print(f"     {name}: {count} cases")
print(f"   P0 (blockers): {sum(1 for m in MODULES.values() for c in m if c[0]=='P0')}")
print(f"   P1 (critical): {sum(1 for m in MODULES.values() for c in m if c[0]=='P1')}")
print(f"   P2 (major):    {sum(1 for m in MODULES.values() for c in m if c[0]=='P2')}")
print(f"   P3 (minor):    {sum(1 for m in MODULES.values() for c in m if c[0]=='P3')}")
